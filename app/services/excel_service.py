"""
Excel Service – generates the Final Excel workbook.

Produces an 8-sheet workbook matching the reference Final Excel format:
  1. Summary          – study metadata + peak labels
  2-5. Class sheets   – cleaned 15-min data per vehicle class
  6. Total Volume Class Breakdown
  7. AM Peak Class Breakdown
  8. PM Peak Class Breakdown

Headers use a merged "Leg — Direction" row to reduce visual clutter.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import PeakResult, StudyMetadata

logger = logging.getLogger(__name__)


class ExcelService:
    """Generates the Final Excel workbook from processed data."""

    # Styling constants
    _HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    _SUBHEADER_FONT = Font(bold=True, size=10)
    _HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    _SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    _ZEBRA_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    def generate(
        self,
        metadata: StudyMetadata,
        class_dfs: dict[str, pd.DataFrame],
        total_df: pd.DataFrame,
        am_breakdown: pd.DataFrame,
        pm_breakdown: pd.DataFrame,
        total_breakdown: pd.DataFrame,
        am_peak: PeakResult | None,
        pm_peak: PeakResult | None,
        output_path: Path,
    ) -> Path:
        """
        Write the Final Excel workbook to *output_path*.

        Returns the path to the created file.
        """
        wb = Workbook()

        # 1. Summary sheet
        self._write_summary(wb.active, metadata, am_peak, pm_peak)
        wb.active.title = "Summary"

        # 2-5. Class sheets
        for class_name in [
            "Lights and Motorcycles",
            "Heavy",
            "Pedestrians",
            "Bicycles on Crosswalk",
        ]:
            ws = wb.create_sheet(title=class_name)
            if class_name in class_dfs:
                self._write_class_sheet(ws, class_dfs[class_name])

        # 6. Total Volume Class Breakdown
        ws_total = wb.create_sheet(title="Total Volume Class Breakdown")
        self._write_breakdown_sheet(ws_total, total_breakdown)

        # 7. AM Peak Class Breakdown
        ws_am = wb.create_sheet(title="AM Peak Class Breakdown")
        self._write_breakdown_sheet(ws_am, am_breakdown)

        # 8. PM Peak Class Breakdown
        ws_pm = wb.create_sheet(title="PM Peak Class Breakdown")
        self._write_breakdown_sheet(ws_pm, pm_breakdown)

        wb.save(output_path)
        logger.info("Final Excel written to %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------

    def _write_summary(self, ws, metadata: StudyMetadata, am_peak, pm_peak):
        rows = [
            ("Study Name", metadata.study_name),
            ("Project", metadata.project),
            ("Project Code", metadata.project_code),
            ("Legs and Movements", metadata.legs_and_movements),
            ("Bin Size", metadata.bin_size),
            ("Time Zone", metadata.timezone),
            ("Start Time", str(metadata.start_time) if metadata.start_time else ""),
            ("End Time", str(metadata.end_time) if metadata.end_time else ""),
            ("Location", metadata.location),
            (
                "Latitude and Longitude",
                f"{metadata.latitude} , {metadata.longitude}"
                if metadata.latitude
                else "",
            ),
            ("", ""),
            ("AM Peak", am_peak.label if am_peak else ""),
            (
                "PM Peak (Overall Peak Hour)",
                pm_peak.label if pm_peak else "",
            ),
        ]
        for r, (key, val) in enumerate(rows, start=1):
            ws.cell(row=r, column=1, value=key).font = Font(bold=True, size=11)
            ws.cell(row=r, column=2, value=val)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 50

    # ------------------------------------------------------------------
    # Class data sheet (merged Leg — Direction header)
    # ------------------------------------------------------------------

    def _write_class_sheet(self, ws, df: pd.DataFrame):
        """Write a per-class DataFrame with merged Leg–Direction header."""
        if df.empty:
            return

        legs = [c[0] for c in df.columns]
        dirs = [c[1] for c in df.columns]
        movs = [c[2] for c in df.columns]

        # Row 1: Merged "Leg — Direction"
        ws.cell(row=1, column=1, value="Approach").font = self._HEADER_FONT
        ws.cell(row=1, column=1).fill = self._HEADER_FILL
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Row 2: Movement
        ws.cell(row=2, column=1, value="Start Time").font = self._SUBHEADER_FONT
        ws.cell(row=2, column=1).fill = self._SUBHEADER_FILL
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        # Build merged header groups
        col = 2
        for c_idx, (leg, direction, mov) in enumerate(zip(legs, dirs, movs), start=2):
            # Row 1: "Leg — Direction" (merged across same approach)
            approach_label = f"{leg} — {direction}" if leg and direction else leg or direction
            cell1 = ws.cell(row=1, column=c_idx, value=approach_label)
            cell1.font = self._HEADER_FONT
            cell1.fill = self._HEADER_FILL
            cell1.alignment = Alignment(horizontal="center")

            # Row 2: Movement name
            cell2 = ws.cell(row=2, column=c_idx, value=mov)
            cell2.font = self._SUBHEADER_FONT
            cell2.fill = self._SUBHEADER_FILL
            cell2.alignment = Alignment(horizontal="center")

        # Merge cells in row 1 for same approach
        self._merge_approach_cells(ws, legs, dirs, row=1, start_col=2)

        # Data rows
        for r, (ts, row_data) in enumerate(df.iterrows(), start=3):
            ws.cell(row=r, column=1, value=ts)
            for c, val in enumerate(row_data, start=2):
                cell = ws.cell(row=r, column=c, value=int(val) if pd.notna(val) else 0)
                cell.alignment = Alignment(horizontal="center")
                if r % 2 == 1:
                    cell.fill = self._ZEBRA_FILL

        # Apply borders
        self._apply_borders(ws, max_row=len(df) + 2, max_col=len(df.columns) + 1)

    # ------------------------------------------------------------------
    # Breakdown sheet (merged Leg — Direction header)
    # ------------------------------------------------------------------

    def _write_breakdown_sheet(self, ws, df: pd.DataFrame):
        """Write a breakdown DataFrame with merged Leg–Direction header."""
        if df.empty:
            return

        cols = list(df.columns)
        legs = [c[0] if isinstance(c, tuple) else "" for c in cols]
        dirs = [c[1] if isinstance(c, tuple) else "" for c in cols]
        movs = [c[2] if isinstance(c, tuple) else str(c) for c in cols]

        # Row 1: Merged "Leg — Direction"
        ws.cell(row=1, column=1, value="Approach").font = self._HEADER_FONT
        ws.cell(row=1, column=1).fill = self._HEADER_FILL
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Row 2: Movement
        ws.cell(row=2, column=1, value="Start Time").font = self._SUBHEADER_FONT
        ws.cell(row=2, column=1).fill = self._SUBHEADER_FILL
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        for c_idx, (leg, direction, mov) in enumerate(zip(legs, dirs, movs), start=2):
            approach_label = f"{leg} — {direction}" if leg and direction else leg or direction
            cell1 = ws.cell(row=1, column=c_idx, value=approach_label)
            cell1.font = self._HEADER_FONT
            cell1.fill = self._HEADER_FILL
            cell1.alignment = Alignment(horizontal="center")

            cell2 = ws.cell(row=2, column=c_idx, value=mov)
            cell2.font = self._SUBHEADER_FONT
            cell2.fill = self._SUBHEADER_FILL
            cell2.alignment = Alignment(horizontal="center")

        # Merge cells in row 1 for same approach
        self._merge_approach_cells(ws, legs, dirs, row=1, start_col=2)

        # Data rows
        for r, (idx, row_data) in enumerate(df.iterrows(), start=3):
            label = idx
            if hasattr(idx, "strftime"):
                label = idx
            ws.cell(row=r, column=1, value=label)

            for c, val in enumerate(row_data, start=2):
                if pd.isna(val):
                    cell = ws.cell(row=r, column=c, value="")
                elif isinstance(val, float) and val != int(val):
                    cell = ws.cell(row=r, column=c, value=val)
                else:
                    try:
                        cell = ws.cell(row=r, column=c, value=int(val))
                    except (ValueError, TypeError):
                        cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(horizontal="center")
                if r % 2 == 1:
                    cell.fill = self._ZEBRA_FILL

        # Apply borders
        self._apply_borders(ws, max_row=len(df) + 2, max_col=len(cols) + 1)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _merge_approach_cells(self, ws, legs, dirs, row, start_col):
        """Merge adjacent cells in a row that share the same Leg+Direction."""
        if not legs:
            return

        merge_start = start_col
        prev_key = (legs[0], dirs[0])

        for i in range(1, len(legs)):
            curr_key = (legs[i], dirs[i])
            if curr_key != prev_key:
                # Merge the previous group
                if merge_start < start_col + i:
                    ws.merge_cells(
                        start_row=row, start_column=merge_start,
                        end_row=row, end_column=start_col + i - 1,
                    )
                merge_start = start_col + i
                prev_key = curr_key

        # Merge the last group
        if merge_start < start_col + len(legs):
            ws.merge_cells(
                start_row=row, start_column=merge_start,
                end_row=row, end_column=start_col + len(legs) - 1,
            )

    def _apply_borders(self, ws, max_row, max_col):
        """Apply thin borders to all cells in the data range."""
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = self._THIN_BORDER
